# -*- coding: utf-8 -*-

__author__ = 'jingqiwang'


from bs4 import BeautifulSoup
from openpyxl import load_workbook
import urllib2, sys, re

reload(sys)
sys.setdefaultencoding('utf-8')


class MovieInfo:
    def __init__(self, title, rating, vote, link, year, director, screenwriter, actors, genres, country, language, release_date, duration, episodes):
        self.title = title
        self.rating = rating
        self.vote = vote
        self.link = link
        self.year = year
        self.director = director
        self.screenwriter = screenwriter
        self.actors = actors
        self.genres = genres
        self.country = country
        self.language = language
        self.release_date = release_date
        self.duration = duration
        self.episodes = episodes


class Spider:
    def __init__(self):
        pass

    def crawler(self, tag):
        page_num = 0
        wb = load_workbook('movie.xlsx')
        # ws = wb.active
        ws = wb.create_sheet(tag)
        ws.append(['名称', '年份', '评分', '评分人数', '链接', '导演', '编剧', '演员', '类型', '国家/地区', '语言', '上映日期', '片长', '集数'])
        '''
        add_info = 'insert into romance_movie values (null, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'
        db = '/Users/jingqiwang/PycharmProjects/Crawler/douban/Movie/movie.db'
        conn = sqlite3.connect(db)
        cursor = conn.cursor()
        '''
        while True:
            url = 'http://www.douban.com/tag/%s/movie?start=%s' % (tag, str(page_num * 15))
            headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux i686; rv:8.0) Gecko/20100101 Firefox/8.0 Chrome/20.0.1132.57 Safari/536.11'}
            try:
                request = urllib2.Request(url, headers=headers)
                response = urllib2.urlopen(request)
                page = response.read()
            except Exception, e:
                print e
                sys.exit(1)

            soup = BeautifulSoup(page, 'html.parser')
            if soup.find_all('dd'):
                for item in soup.find_all('dd'):
                    title = item.find('a').text
                    link = item.find('a')['href'].split('?')[0]
                    try:
                        rating = item.find('span', class_='rating_nums').text
                    except Exception:
                        rating = '--'
                        print link, 'rating'

                    try:
                        request = urllib2.Request(link, headers=headers)
                        response = urllib2.urlopen(request)
                        page = response.read()
                    except Exception:
                        print link
                        continue

                    detail = BeautifulSoup(page, 'html.parser')
                    try:
                        year = detail.find('span', class_='year').text[1:-1]
                    except Exception:
                        year = '--'
                        print link, 'year'

                    try:
                        director = detail.find('a', rel='v:directedBy').text
                    except Exception:
                        director = '--'
                        print link, 'director'

                    try:
                        screenwriter = detail.find('span', text='编剧').next_sibling.next_sibling.text
                    except Exception:
                        screenwriter = '--'
                        print link, 'screenwriter'

                    try:
                        actors = ' / '.join([span.text for span in detail.find('span', class_='actor').find_all('a', rel='v:starring')])
                    except Exception:
                        actors = '--'
                        print link, 'actors'

                    try:
                        genres = ' / '.join([span.text for span in detail.find_all('span', property='v:genre')])
                    except Exception:
                        genres = '--'
                        print link, 'genres'

                    try:
                        country = detail.find('span', text='制片国家/地区:').next_sibling
                    except Exception:
                        country = '--'
                        print link, 'country'

                    try:
                        language = detail.find('span', text='语言:').next_sibling
                    except Exception:
                        language = '--'
                        print link, 'language'

                    try:
                        release_date = ' / '.join([span.text for span in detail.find_all('span', property='v:initialReleaseDate')])
                    except Exception:
                        release_date = '--'
                        print link, 'release_date'

                    try:
                        duration = detail.find('span', property='v:runtime').text
                    except Exception:
                        duration = '--'
                        print link, 'duration'

                    try:
                        episodes = detail.find('span', text='集数:').next_sibling
                    except Exception:
                        episodes = '--'

                    try:
                        vote = detail.find('span', property='v:votes').text
                    except Exception:
                        vote = '--'
                        print link, 'vote'
                    '''
                    info = (title, rating, vote, link, year, director, screenwriter, actors, genres, country, language, release_date, duration, episodes)
                    cursor.execute(add_info, info)
                    conn.commit()
                    '''
                    try:
                        ws.append([title, year, rating, vote, link, director, screenwriter, actors, genres, country, language, release_date, duration, episodes])
                    except Exception:
                        title = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', title)
                        year = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', year)
                        rating = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', rating)
                        vote = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', vote)
                        link = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', link)
                        director = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', director)
                        screenwriter = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', screenwriter)
                        actors = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', actors)
                        genres = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', genres)
                        country = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', country)
                        language = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', language)
                        release_date = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', release_date)
                        duration = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', duration)
                        episodes = re.sub('[\x00-\x08]|[\x0b-\x0c]|[\x0e-\x1f]', '', episodes)
                        ws.append([title, year, rating, vote, link, director, screenwriter, actors, genres, country, language, release_date, duration, episodes])
            else:
                '''
                cursor.close()
                conn.close()
                '''
                wb.save('movie.xlsx')
                break

            page_num += 1

m = Spider()
m.crawler(u'文艺')
